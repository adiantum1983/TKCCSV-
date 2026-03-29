import streamlit as st  # type: ignore
import pandas as pd  # type: ignore
import io
from cashflow_direct import load_trial_balance, create_direct_cf_statement  # type: ignore
from financial_metrics import compute_financial_metrics

st.set_page_config(page_title="財務分析アプリ", layout="centered")

st.title("📊 財務分析アプリ")
st.markdown("お手元の **残高試算表 (ExcelまたはCSV)** をアップロードするだけで、簡易な直接法のキャッシュフロー計算書の作成と財務指標の分析を自動で行います。")

uploaded_file = st.file_uploader("残高試算表のファイル（Excel/CSV）をドラッグ＆ドロップ（または選択）", type=["xlsx", "xls", "csv"])

if uploaded_file is not None:
    st.info("ファイルを読み込み、マトリックス法による計算を行っています...")
    try:
        # ファイルパスの代わりに BytesIO のラッパーを渡しても pandas.read_excel はパース可能
        tb_df = load_trial_balance(uploaded_file)
        df_cf = create_direct_cf_statement(tb_df)
        
        st.success("計算完了！以下のタブを切り替えて各レポートを確認できます。")
        
        tab1, tab2 = st.tabs(["📄 キャッシュフロー計算書", "📈 財務指標分析 (月次・四半期)"])
        
        with tab1:
            st.subheader("生成された直接法キャッシュフロー計算書")
            st.dataframe(df_cf, use_container_width=True, hide_index=True)
            
        with tab2:
            st.subheader("財務指標分析レポート")
            
            # 指標計算関数の呼び出し（同じファイルを使うのでseekで戻す）
            if hasattr(uploaded_file, "seek"):
                uploaded_file.seek(0)
            df_monthly, df_quarterly = compute_financial_metrics(uploaded_file)
            
            st.markdown("##### 🔍 月次指標 (異常検知用)")
            st.dataframe(df_monthly, use_container_width=True, hide_index=True)
            
            st.markdown("##### 📊 四半期指標 (体質分析用)")
            st.dataframe(df_quarterly, use_container_width=True, hide_index=True)
        
        # ダウンロード用のExcel生成
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df_cf.to_excel(writer, index=False, sheet_name="キャッシュフロー計算書", header=["項目", "金額"])
            worksheet = writer.sheets["キャッシュフロー計算書"]
            
            from openpyxl.styles import Font, Border, Side, PatternFill  # type: ignore
            thin = Side(border_style="thin", color="000000")
            header_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
            
            worksheet.column_dimensions['A'].width = 50
            worksheet.column_dimensions['B'].width = 20
            
            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=2):
                for cell in row:
                    if cell.row == 1:
                        cell.fill = header_fill
                        cell.font = Font(bold=True)  # type: ignore
                    elif cell.col_idx == 2 and isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0'
                    if isinstance(row[0].value, str):  # type: ignore
                        val = row[0].value  # type: ignore
                        if "I." in val or "II." in val or "III." in val or "小計" in val or "増減額" in val or "期末残高" in val:
                             row[0].font = Font(bold=True)  # type: ignore
                             row[1].font = Font(bold=True)  # type: ignore
                             
        st.download_button(
            label="⬇️ Excel形式でダウンロード",
            data=buffer.getvalue(),
            file_name="CashFlow_DirectStatement.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
    except Exception as e:
        import traceback
        st.error(f"データの処理中にエラーが発生しました: {str(e)}")
        with st.expander("エラー詳細 (開発者向け)"):
            st.text(traceback.format_exc())
            st.write("※A列に「勘定科目コード」(1101, 1122等) が正しく付与された残高試算表かどうか再度ご確認ください。")
