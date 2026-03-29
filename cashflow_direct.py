import pandas as pd  # type: ignore
import argparse
import os
import subprocess

def load_trial_balance(filepath):
    filename = filepath if isinstance(filepath, str) else getattr(filepath, "name", "unknown")
    is_csv = filename.lower().endswith(".csv")

    print(f"[{filename}] を読み込んでいます...")
    try:
        if is_csv:
            try:
                df = pd.read_csv(filepath, header=None, encoding="utf-8")
            except UnicodeDecodeError:
                if hasattr(filepath, "seek"): filepath.seek(0)  # type: ignore
                df = pd.read_csv(filepath, header=None, encoding="shift_jis")
        else:
            df = pd.read_excel(filepath, header=None)
    except Exception as e:
        raise RuntimeError(f"ファイルの読み込みに失敗しました: {e}")

    header_idx = -1
    for i, row in df.iterrows():
        if any(isinstance(val, str) and ("科目" in val or "勘定" in val) for val in row.values):
            header_idx = int(str(i)) if str(i).isdigit() else i
            break
            
    if isinstance(header_idx, int) and header_idx > 0:
        if hasattr(filepath, "seek"): filepath.seek(0)  # type: ignore
        
        if is_csv:
            try:
                df = pd.read_csv(filepath, header=header_idx, encoding="utf-8")
            except UnicodeDecodeError:
                if hasattr(filepath, "seek"): filepath.seek(0)  # type: ignore
                df = pd.read_csv(filepath, header=header_idx, encoding="shift_jis")
        else:
            df = pd.read_excel(filepath, header=header_idx)

    df.rename(columns={df.columns[0]: "Code"}, inplace=True)
    df["Code"] = df["Code"].astype(str).str.strip()
    
    df.rename(columns={
        df.columns[1]: "Name",
        df.columns[2]: "Prev",
        df.columns[3]: "Dr",
        df.columns[4]: "Cr",
        df.columns[5]: "Curr"
    }, inplace=True)
    
    numeric_cols = ["Prev", "Dr", "Cr", "Curr"]
    for c in numeric_cols:
         df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)
    
    return df

def get_row_val(df, code, col):
    matched = df[df["Code"] == str(code)]
    if not matched.empty:
        return matched.iloc[0][col]
    return 0

def create_direct_cf_statement(df):
    """
    マトリックス法の貸借一致ロジックを利用し、直接法キャッシュフロー計算書を作成する。
    """
    
    # ======== BSの当期増減(Net Change) ========
    # 資産（Curr-Prev）
    nc_1101 = get_row_val(df, "1101", "Curr") - get_row_val(df, "1101", "Prev")  # 現預金
    nc_1122 = get_row_val(df, "1122", "Curr") - get_row_val(df, "1122", "Prev")  # 売掛金
    nc_1120 = get_row_val(df, "1120", "Curr") - get_row_val(df, "1120", "Prev")  # 棚卸資産
    nc_1130 = get_row_val(df, "1130", "Curr") - get_row_val(df, "1130", "Prev")  # その他流動資産
    nc_1200 = get_row_val(df, "1200", "Curr") - get_row_val(df, "1200", "Prev")  # 固定資産
    
    # 負債・純資産（Curr-Prev）
    nc_2112 = get_row_val(df, "2112", "Curr") - get_row_val(df, "2112", "Prev")  # 買掛金
    nc_2113 = get_row_val(df, "2113", "Curr") - get_row_val(df, "2113", "Prev")  # 短期借入金
    
    # 2100(流動負債計)から買掛金(2112)と短借(2113)を引いたものを「その他流動負債」とする
    nc_2100 = get_row_val(df, "2100", "Curr") - get_row_val(df, "2100", "Prev")
    nc_other_cl = nc_2100 - nc_2112 - nc_2113
    
    nc_2200 = get_row_val(df, "2200", "Curr") - get_row_val(df, "2200", "Prev")  # 固定負債(長期借入金含む)
    nc_3000 = get_row_val(df, "3000", "Curr") - get_row_val(df, "3000", "Prev")  # 純資産

    # ======== PLの発生額 ========
    sales_amount = get_row_val(df, "4000", "Cr") - get_row_val(df, "4000", "Dr")  # 4000 売上高
    cogs_amount  = get_row_val(df, "5200", "Dr") - get_row_val(df, "5200", "Cr")  # 5200 売上原価
    sga_amount   = get_row_val(df, "6100", "Dr") - get_row_val(df, "6100", "Cr")  # 6100 販管費
    
    # 減価償却費 (6214, 5455)
    depr_6214 = get_row_val(df, "6214", "Dr") - get_row_val(df, "6214", "Cr")
    depr_5455 = get_row_val(df, "5455", "Dr") - get_row_val(df, "5455", "Cr")
    depr_total = depr_6214 + depr_5455

    # ======== マトリックス法による各キャッシュフローの算定 ========
    
    # 現金・売掛金・その他流動・棚卸・固定 ＝ 買掛金・短借・その他流動負債・固定負債・純資産
    # ΔCash = -ΔAR -ΔInv -ΔOtherCA -ΔFA + ΔAP + ΔST_Loan + ΔOtherCL + ΔFixedLiab + ΔEquity
    
    # 1. 営業収入マトリックス
    cfo_sales = sales_amount - nc_1122

    # 2. 仕入支出マトリックス
    cfo_cogs = -cogs_amount - nc_1120 + nc_2112

    # 3. 販管費支出マトリックス (減価償却費は非現金のため除く)
    cfo_sga = -(sga_amount - depr_total) + nc_other_cl - nc_1130

    # 4. その他営業・営業外損益マトリックス
    # 純資産の変動から、売上、原価、販管費（本業PL）を差し引いた残りが営業外損益や税金・配当等
    cfo_other = nc_3000 - (sales_amount - cogs_amount - sga_amount)
    
    cfo_total = cfo_sales + cfo_cogs + cfo_sga + cfo_other

    # 5. 投資活動マトリックス
    # 減価償却による固定資産減少を現金のマイナスから排除
    cfi_total = -nc_1200 - depr_total

    # 6. 財務活動マトリックス
    cff_total = nc_2200 + nc_2113

    # ======== CFレポートフォーマット作成 ========
    import typing
    rows: typing.List[typing.Dict[str, typing.Any]] = []
    
    rows.append({"Category": "I. 営業活動によるキャッシュ・フロー", "Amount": ""})
    rows.append({"Category": "　営業収入", "Amount": int(cfo_sales)})
    rows.append({"Category": "　商品の仕入れによる支出", "Amount": int(cfo_cogs)})
    rows.append({"Category": "　人件費・その他営業活動による支出", "Amount": int(cfo_sga)})
    rows.append({"Category": "　営業外活動・税金等による増減", "Amount": int(cfo_other)})
    rows.append({"Category": "営業活動によるキャッシュ・フロー（小計）", "Amount": int(cfo_total)})
    rows.append({"Category": "", "Amount": ""})
    
    rows.append({"Category": "II. 投資活動によるキャッシュ・フロー", "Amount": ""})
    rows.append({"Category": "　有形・無形固定資産等への投資活動", "Amount": int(cfi_total)})
    rows.append({"Category": "投資活動によるキャッシュ・フロー（小計）", "Amount": int(cfi_total)})
    rows.append({"Category": "", "Amount": ""})

    rows.append({"Category": "III. 財務活動によるキャッシュ・フロー", "Amount": ""})
    rows.append({"Category": "　借入金の増減等", "Amount": int(cff_total)})
    rows.append({"Category": "財務活動によるキャッシュ・フロー（小計）", "Amount": int(cff_total)})
    rows.append({"Category": "", "Amount": ""})

    calc_total_cf = cfo_total + cfi_total + cff_total
    
    rows.append({"Category": "現金及び現金同等物の増減額", "Amount": int(calc_total_cf)})
    rows.append({"Category": "現金及び現金同等物の期首残高", "Amount": int(get_row_val(df, "1101", "Prev"))})
    rows.append({"Category": "現金及び現金同等物の期末残高 (計算値)", "Amount": int(get_row_val(df, "1101", "Prev") + calc_total_cf)})
    rows.append({"Category": "現金及び現金同等物の期末残高 (実際値)", "Amount": int(get_row_val(df, "1101", "Curr"))})
    
    # 差額チェック
    diff = int(get_row_val(df, "1101", "Curr")) - int(get_row_val(df, "1101", "Prev") + calc_total_cf)
    if diff != 0:
        rows.append({"Category": "【警告】貸借差額エラー", "Amount": diff})

    return pd.DataFrame(rows)

def generate_report(df_cf, output_path):
    import openpyxl  # type: ignore
    from openpyxl.styles import Font, Border, Side, PatternFill
    
    df_cf.to_excel(output_path, index=False, sheet_name="キャッシュフロー計算書", header=["項目", "金額"])
    
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active
    
    thin = Side(border_style="thin", color="000000")  # type: ignore
    border = Border(top=thin, bottom=thin)  # type: ignore
    header_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")  # type: ignore
    
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 20
    
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=2):
        for cell in row:
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = Font(bold=True)  # type: ignore
            elif cell.col_idx == 2 and isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0'
                
            if "I." in str(row[0].value) or "II." in str(row[0].value) or "III." in str(row[0].value) or "小計" in str(row[0].value) or "増減額" in str(row[0].value) or "期末残高" in str(row[0].value):  # type: ignore
                 row[0].font = Font(bold=True)  # type: ignore
                 row[1].font = Font(bold=True)  # type: ignore

    wb.save(output_path)
    print(f"成功: Excelレポートを '{output_path}' に保存しました。")

def convert_to_pdf(excel_path, pdf_path):
    print(f"PDFへ変換中: {pdf_path}")
    ps_script = f"""
    $excel = New-Object -ComObject Excel.Application
    $excel.Visible = $false
    $excel.DisplayAlerts = $false
    try {{
        $workbook = $excel.Workbooks.Open("{excel_path}")
        $sheet = $workbook.Sheets.Item(1)
        
        # 印刷設定を見やすくする
        $sheet.PageSetup.Zoom = $false
        $sheet.PageSetup.FitToPagesWide = 1
        $sheet.PageSetup.FitToPagesTall = 1
        $sheet.PageSetup.CenterHorizontally = $true
        
        $workbook.ExportAsFixedFormat(0, "{pdf_path}")
        $workbook.Close($false)
        Write-Host "PDF変換が完了しました"
    }} catch {{
        Write-Host "エラーが発生しました: $_"
    }} finally {{
        $excel.Quit()
        [System.Runtime.Interopservices.Marshal]::ReleaseComObject($excel) | Out-Null
    }}
    """
    script_path = os.path.join(os.path.dirname(excel_path), "export_pdf.ps1")
    with open(script_path, "w", encoding="utf-8-sig") as f:
        f.write(ps_script)
    
    subprocess.run(["powershell", "-ExecutionPolicy", "Bypass", "-File", script_path])
    os.remove(script_path)

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("input_file")
    parser.add_argument("-o", "--output", default="CashFlow_DirectStatement.xlsx")
    args = parser.parse_args()

    input_path = os.path.abspath(args.input_file)
    output_excel = os.path.abspath(args.output)
    output_pdf = output_excel.replace(".xlsx", ".pdf")

    if not os.path.exists(input_path):
        print(f"エラー: 指定された入力ファイル '{input_path}' が見つかりません。")
        return

    tb_df = load_trial_balance(input_path)
    df_cf = create_direct_cf_statement(tb_df)
    generate_report(df_cf, output_excel)
    
    convert_to_pdf(output_excel, output_pdf)

if __name__ == "__main__":
    main()
